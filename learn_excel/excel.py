# coding:utf-8

import traceback
import xlsxwriter
import string
from io import BytesIO
from xlsxwriter import Workbook as writer_book
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Color, Font, Alignment, PatternFill, Border, Side
from tornado.log import app_log
from web import utils
letter_list = string.uppercase

class ExcelStyles(object):
    TOTAL_FONT = Font(color="FFFFFFFF", size=16.0)
    TOTAL_FG_COLOR = Color(rgb="FF558ED5")
    TOTAL_BG_COLOR = Color(rgb="FF808080")
    BG_COLOR2 = Color(rgb="FFD7E4BD")
    BG_COLOR3 = Color(rgb="FFD7E4BD")
    FG_COLOR3 = Color(rgb="FFC3D69B")
    FG_COLOR2 = Color(rgb="FF93CDDD")
    # 列名颜色
    FG_FILL = PatternFill("solid", fgColor=FG_COLOR2, bgColor=BG_COLOR2)
    # 利用率颜色
    FG_FILL1 = PatternFill("solid", fgColor=FG_COLOR3, bgColor=BG_COLOR3)

    FONT2 = Font(color="FF000000", size=12,
                 name="Droid Sans Fallback", family=2.0)
    TOTAL_FILL = PatternFill("solid", fgColor=TOTAL_FG_COLOR,
                             bgColor=TOTAL_BG_COLOR)
    ALIGNMENT = Alignment(vertical="center", horizontal="center", wrapText=True)
    SIDE = Side(style="thin")
    BORDER_SIDE = "thin"
    BORDER = Border(left=SIDE, right=SIDE, outline=True, top=SIDE, bottom=SIDE)
    FONT_COLOR = "black"
    FONT_COLOR2 = "#FF000000"
    TOTAL_FONT_COLOR = "FFFFFFFF"
    TOTAL_FONT_SIZE = 16.0
    TOTAL_FG_COLOR_SET = "#FF558ED5"
    TOTAL_BG_COLOR_SET = "#558ED5"
    FONT_SIZE = 12
    FONT_NAME = "Droid Sans Fallback"
    BG_COLOR_SET = "#C3D69B"
    BG_COLOR_SET1 = "#93CDDD"
    FG_COLOR_SET = "FFC3D69B"
    FG_COLOR_SET1 = "FG_COLOR2"
    DEFAULT_FORMAT = {"font_name": FONT_NAME, "font_color": FONT_COLOR,
                      "font_size": FONT_SIZE, "border": True, "align": "center",
                      "valign": "vcenter"}
    J_ALARM_COLOR1 = "00C6E0B4"
    J_ALARM_COLOR2 = "00DDEBF7"
    J_ALARM_FILL1 = PatternFill("solid", fgColor=J_ALARM_COLOR1, bgColor=BG_COLOR2)
    J_ALARM_FILL2 = PatternFill("solid", fgColor=J_ALARM_COLOR2, bgColor=BG_COLOR2)
    J_ALARM_FONT2 = Font(size=11, name=u"等线")
    J_ALARM_FONT1 = Font(name=u"等线")
    J_ALARM_FONT0 = Font(size=14, name=u"等线")


class ExcelService(object):
    def __init__(self):
        self.font = ExcelStyles.FONT2

    def create_sheet(self, work, name, merges=list(), add_data=dict(), index=-1):
        sheet = work.create_sheet(title=name, index=index)
        for merge in merges:
            sheet.merge_cells(merge)
        for k, v in add_data.items():
            sheet[k] = v
            sheet[k].alignment = ExcelStyles.ALIGNMENT
        return sheet

    def add_excel_cell(self, sheet, all_data, row_num=0, order_list=None,
                       font=ExcelStyles.FONT2, height=25.0, fill=PatternFill(),
                       border=ExcelStyles.BORDER, start_col=0, add_num=1):
        sheet.row_dimensions[row_num].height = height
        # app_log.info(all_data)
        if not border:
            border = Border()
        if isinstance(all_data, list):
            if isinstance(all_data[0], dict):
                for data in all_data:
                    for index, value in enumerate(order_list):
                        utils.encode_unicode(value)
                        place = letter_list[index+start_col] \
                                + str(row_num)
                        try:
                            sheet[place] = data[value]
                        except Exception as e:
                            app_log.error(traceback.format_exc())
                            pass
                        sheet[place].font = font
                        sheet[place].alignment = ExcelStyles.ALIGNMENT
                        sheet[place].border = border
                        sheet[place].fill = fill
                    row_num += add_num
            elif isinstance(all_data[0], tuple):
                for value in all_data:
                    utils.encode_unicode(value)
                    place = value[0]
                    try:
                        sheet[place] = value[1]
                    except Exception as e:
                        app_log.error(traceback.format_exc())
                        pass
                    sheet[place].font = font
                    sheet[place].alignment = ExcelStyles.ALIGNMENT
                    sheet[place].border = border
                    sheet[place].fill = fill
                row_num += add_num
            elif isinstance(all_data[0], list):
                for info in all_data:
                        for index, value in enumerate(info):
                            utils.encode_unicode(value)
                            place = letter_list[
                                        index + start_col] + \
                                    str(row_num)
                            try:
                                sheet[place] = value
                            except Exception as e:
                                app_log.error(traceback.format_exc())
                                pass
                            sheet[place].font = font
                            sheet[place].alignment = ExcelStyles.ALIGNMENT
                            sheet[place].border = border
                            sheet[place].fill = fill
                        row_num += add_num

            else:
                for index, value in enumerate(all_data):
                    place = letter_list[index + start_col] +\
                            str(row_num)
                    # app_log.info([place, value])
                    utils.encode_unicode(value)
                    try:
                        sheet[place] = value
                    except Exception as e:
                        app_log.error(traceback.format_exc())
                        pass
                    sheet[place].font = font
                    sheet[place].alignment = ExcelStyles.ALIGNMENT
                    sheet[place].border = border
                    sheet[place].fill = fill
                row_num += add_num
        else:
            for index, value in enumerate(order_list):
                utils.encode_unicode(value)
                place = letter_list[index + start_col] + \
                        str(row_num)
                try:
                    sheet[place] = all_data[value]
                except Exception as e:
                    app_log.error(traceback.format_exc())
                    pass
                sheet[place].font = font
                sheet[place].alignment = ExcelStyles.ALIGNMENT
                sheet[place].border = border
                sheet[place].fill = fill
            row_num += add_num
        return row_num

    def build_format(self, work_book, bg_color="", fg_color="", align="center",
                     font_name=ExcelStyles.FONT_NAME, valign="vcenter",
                     font_color=ExcelStyles.FONT_COLOR,
                     font_size=ExcelStyles.FONT_SIZE
                     ):
        cell_format = {"font_name": font_name, "font_color": font_color,
                       "font_size": font_size, "border": True,
                       "align": align, "valign": valign,
                       "bg_color": bg_color,
                       "fg_color": fg_color}
        cell_format = work_book.add_format(cell_format)
        return cell_format

    def new_add_excel_cell(self, sheet, all_data, align="center",
                           font_name=ExcelStyles.FONT_NAME, valign="vcenter",
                           font_color= ExcelStyles.FONT_COLOR, row_num=0,
                           font_size=ExcelStyles.FONT_SIZE, order_list=None,
                           height=25.0, start_col=0, add_num=1, bg_color="",
                           fg_color=""):
        cell_format = {"font_name": font_name, "font_color": font_color,
                          "font_size": font_size, "border": True,
                          "align": align, "valign": valign,
                          "bg_color": bg_color,
                          "fg_color": fg_color}
        cell_format = xlsxwriter.Workbook
        sheet.set_row(row_num, height, cell_format)
        app_log.info(all_data)
        if isinstance(all_data, list):
            if isinstance(all_data[0], dict):
                for data in all_data:
                    for index, value in enumerate(order_list):
                        value = data[value]
                        utils.encode_unicode(value)

                        try:
                            sheet.write(row_num, index + start_col, value,
                                        cell_format)
                        except Exception as e:
                            app_log.error(traceback.format_exc())
                            pass
                    row_num += add_num
            elif isinstance(all_data[0], tuple):
                for value in all_data:
                    place = value[0]
                    try:
                        sheet.write(place, value[1], cell_format)
                    except Exception as e:
                        app_log.error(traceback.format_exc())
                        pass
                row_num += add_num
            elif isinstance(all_data[0], list):
                for info in all_data:
                        for index, value in enumerate(info):
                            utils.encode_unicode(value)
                            try:
                                sheet.write(row_num, index + start_col, value,
                                            cell_format)
                            except Exception as e:
                                app_log.error(traceback.format_exc())
                                pass
                        row_num += add_num

            else:
                for index, value in enumerate(all_data):
                    place = letter_list[index + start_col] +\
                            str(row_num)
                    app_log.info([place, value])
                    utils.encode_unicode(value)
                    try:
                        sheet.write(row_num, index + start_col, value,
                                    cell_format)
                    except Exception as e:
                        app_log.error(traceback.format_exc())
                        pass
                row_num += add_num
        else:
            for index, value in enumerate(order_list):
                utils.encode_unicode(value)
                try:
                    sheet.write(row_num, index + start_col, value,
                                cell_format)
                except Exception as e:
                    app_log.error(traceback.format_exc())
                    pass
            row_num += add_num
        return row_num

    def add_one_row_by_dict(self, sheet, all_data, row_num=0, order_list=None,
                       font=None, height=25.0, fill=None,
                       border=ExcelStyles.BORDER):
        for index, value in enumerate(order_list):
            place = letter_list[index] + str(index + 1)
            sheet[place] = all_data.get(value)
        row_num += 1

    def add_one_row_by_tupe(self, data, sheet, row_num):
        for value in data:
            place = value[0]
            sheet[place] = value[1]
        row_num += 1

    def merge_cells(self, sheet, data):
        for merge in data:
            sheet.merge_cells(merge)

    def unmerge_cells(self, sheet, data):
        for merge in data:
            sheet.unmerge_cells(merge)

    def add_one_cell(self, sheet, all_data, font=ExcelStyles.FONT2,
                     fill=PatternFill(), border=ExcelStyles.BORDER):
        place = all_data[0]
        sheet[place] = all_data[1]
        sheet[place].font = font
        sheet[place].alignment = ExcelStyles.ALIGNMENT
        sheet[place].border = border
        sheet[place].fill = fill

    def set_col_wight(self, sheet, data, wight):
        for place in data:
            sheet.column_dimensions[place].width = wight

    def merge_row(self, sheet, data, start_row):
        for merge in data:
            merge = merge.split(":")
            for index, value in enumerate(merge):
                merge[index] = value + str(start_row)
            merge = ":".join(merge)
            sheet.merge_cells(merge)


class XLSXWRITER_SERVICE():
    def __init__(self):
        self.font = ExcelStyles.FONT2

    def create_sheet(self, work, name, merges=list(), add_data=dict(), index=-1):
        sheet = work.create_sheet(title=name, index=index)
        for merge in merges:
            sheet.merge_cells(merge)
        for k, v in add_data.items():
            sheet[k] = v
            sheet[k].alignment = ExcelStyles.ALIGNMENT
        return sheet

    def build_format(self, work_book, bg_color="#FFFFFF", fg_color="#FFFFFF", align="center",
                     font_name=ExcelStyles.FONT_NAME, valign="vcenter",
                     font_color=ExcelStyles.FONT_COLOR,
                     font_size=ExcelStyles.FONT_SIZE, other_format=None
                     ):
        cell_format = {"font_name": font_name, "font_color": font_color,
                       "font_size": font_size, "border": True,
                       "align": align, "valign": valign, "pattern": 1,
                       "text_wrap": True, "fg_color": fg_color,
                       "bg_color": bg_color}
        if other_format:
            cell_format.update(other_format)
        cell_format = work_book.add_format(cell_format)
        return cell_format

    def add_excel_cell(self, sheet, all_data, cell_format, row_num=0,
                       order_list=None, height=25.0,
                       start_col=0, add_num=1):
        sheet.set_row(row_num, height, cell_format)
        app_log.info(all_data)
        if isinstance(all_data, list):
            if isinstance(all_data[0], dict):
                for data in all_data:
                    for index, value in enumerate(order_list):
                        value = data[value]
                        utils.encode_unicode(value)

                        try:
                            sheet.write(row_num, index + start_col, value,
                                        cell_format)
                        except Exception as e:
                            app_log.error(traceback.format_exc())
                            pass
                    row_num += add_num
            elif isinstance(all_data[0], tuple):
                for value in all_data:
                    place = value[0]
                    try:
                        sheet.write(place, value[1], cell_format)
                    except Exception as e:
                        app_log.error(traceback.format_exc())
                        pass
                row_num += add_num
            elif isinstance(all_data[0], list):
                for info in all_data:
                        for index, value in enumerate(info):
                            utils.encode_unicode(value)
                            try:
                                sheet.write(row_num, index + start_col, value,
                                            cell_format)
                            except Exception as e:
                                app_log.error(traceback.format_exc())
                                pass
                        row_num += add_num

            else:
                for index, value in enumerate(all_data):
                    place = letter_list[index + start_col] +\
                            str(row_num)
                    app_log.info([place, value])
                    utils.encode_unicode(value)
                    try:
                        sheet.write(row_num, index + start_col, value,
                                    cell_format)
                    except Exception as e:
                        app_log.error(traceback.format_exc())
                        pass
                row_num += add_num
        else:
            for index, value in enumerate(order_list):
                utils.encode_unicode(value)
                try:
                    sheet.write(row_num, index + start_col, value,
                                cell_format)
                except Exception as e:
                    app_log.error(traceback.format_exc())
                    pass
            row_num += add_num
        return row_num

    def merge_cells(self, sheet, data):
        for merge in data:
            sheet.merge_range(merge, "")

    def merge_cell(self, sheet, start_row, start_col, end_row,
                   end_col,  add_data, cell_format):
        if isinstance(add_data, str):
            add_data = unicode(add_data, "utf-8")
        sheet.merge_range(
            start_row, start_col, end_row, end_col, add_data, cell_format)

    def add_one_cell(self, sheet, all_data, cell_format):
        place = all_data[0]
        sheet.write(place, all_data[1], cell_format)

    def write_cell(self, sheet, row, col, value, cell_format):
        if isinstance(value, str):
            value = unicode(value, "utf-8")
        sheet.write(row, col, value, cell_format)

    def set_one_column(self, sheet, start_col, end_col,
                       width=15):
        all_column = letter_list
        start_col = all_column[start_col]
        end_col = all_column[end_col]
        range_width = ":".join([start_col, end_col])
        sheet.set_column(range_width, width)

    def add_col_value(self, sheet, cell_format, add_data, start_row=0,
                      start_col=0, ord_list=None, end_col=None,
                      height=100):
        if ord_list:
            for value in ord_list:
                value = add_data[value]
                if isinstance(value, str):
                    value = unicode(value, "utf-8")
                if end_col:
                    sheet.merge_range(start_row, start_col, start_row, end_col,
                                      value, cell_format)
                else:
                    sheet.write(start_row, start_col, value, cell_format)
                sheet.set_row(start_row, height)
                start_row += 1
        else:
            for value in add_data:
                if isinstance(value, str):
                    value = unicode(value, "utf-8")
                if end_col:
                    sheet.merge_range(start_row, start_col, start_row, end_col,
                                      value, cell_format)
                else:
                    sheet.write(start_row, start_col, value, cell_format)
                sheet.set_row(start_row, height)
                start_row += 1
        return start_row

    def add_row_value(self, sheet, cell_format, add_data, start_row=0,
                      start_col=0, ord_list=None, end_row=None, add_col=1,
                      height=100):
        sheet.set_row(start_row, height)
        if ord_list:
            for value in ord_list:
                next_col = start_col + add_col
                end_col = next_col - 1
                value = add_data[value]
                if isinstance(value, str):
                    value = unicode(value, "utf-8")
                if end_row:
                    sheet.merge_range(start_row, start_col, end_row, end_col,
                                      value, cell_format)
                else:
                    sheet.write(start_row, start_col, value, cell_format)
                start_col = next_col
        else:
            for value in add_data:
                next_col = start_col + add_col
                end_col = next_col - 1
                if isinstance(value, str):
                    value = unicode(value, "utf-8")
                if end_row:
                    sheet.merge_range(start_row, start_col, end_row, end_col,
                                      value, cell_format)
                else:
                    sheet.write(start_row, start_col, value, cell_format)
                start_col = next_col
        return start_col

    def set_row(self, sheet, row, height=100):
        sheet.set_row(row, height)

    def get_col(self, index):
        return letter_list[index]


def create_write_excel():
    sheet_name = "test"
    sheet_name = unicode(sheet_name, "utf-8")
    excel = BytesIO()
    workbook = writer_book(excel)
    write_excel_service = XLSXWRITER_SERVICE()
    worksheet = workbook.add_worksheet(sheet_name)
    default_format = write_excel_service.build_format(workbook)
    default_format2 = write_excel_service.build_format(workbook)
    default_format2.set_num_format(0x09)
    workbook.close()
    excel = excel.getvalue()






