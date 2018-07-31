# encoding: utf-8
import re
import os
import json
import string
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Color, Font, Alignment, PatternFill, Border, Side
base_path = os.path.split(os.path.realpath(__file__))[0].split('moses')[0] + "/moses"
language_json_path = base_path + "/front/mobile/static/lang/lang.json"  # 前端国际化json文件位置
language_excel_path = base_path + "/front/mobile/static/lang/language.xlsx"  # 前端国际化excel文件位置
need_trans_path = base_path + "/front/mobile"   # 前端需要国际化的项目目录
trans_list = ["zh", "en"]
trans_dict = {"zh": "中文", "en": "英文"}


def encode_unicode(data):
    if isinstance(data, unicode):
        return data.encode("utf-8")
    return data


def get_trans_message(trans_file_path, result_set):
    with open(trans_file_path) as f:
        text = f.read()
        regex = re.compile("[d][i][1][8][n][(]['\"](.*?)['\"][)]")
        trans = regex.findall(text)
        for message in trans:
            result_set.add(message)


def get_trans_file(dir_path, result_set):
    files = os.listdir(dir_path)
    for file_name in files:
        file_path = dir_path + "/" + file_name
        if os.path.isdir(file_path):
            get_trans_file(file_path, result_set)
        else:
            get_trans_message(file_path, result_set)


def create_need_trans():
    letter_list = string.uppercase
    need_trans = set()
    get_trans_file(need_trans_path, need_trans)
    if os.path.exists(language_json_path):
        with open(language_json_path) as f:
            trans_before = json.loads(f.read())
    else:
        trans_before = {}
    zh_trans = trans_before.get(trans_list[0], {})
    workbook = Workbook()
    sheet = workbook.create_sheet(title="trans", index=0)
    row = 1
    fill = PatternFill()
    font = Font(color="FF000000", size=12,
                 name="Droid Sans Fallback", family=2.0)
    alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
    side = Side(style="thin")
    border = Border(left=side, right=side, outline=True, top=side, bottom=side)
    row_height = 25.0  # 行高
    col_wight = 100.0  # 列宽

    def add_one_cell(start_row, start_col, add_value):
        sheet.row_dimensions[start_row].height = row_height
        place = letter_list[start_col] + str(start_row)
        sheet[place] = add_value
        sheet[place].font = font
        sheet[place].alignment = alignment
        sheet[place].border = border
        sheet[place].fill = fill
    col = 0
    for index, trans_name in enumerate(trans_list):
        col_name = letter_list[index]
        sheet.column_dimensions[col_name].width = col_wight
        add_one_cell(row, col, trans_dict[trans_name])
        col += 1
    row += 1
    for value in zh_trans.keys():
        col = 0
        value_encode = encode_unicode(value)
        need_trans.remove(value_encode)
        add_one_cell(row, col, value_encode)
        col += 1
        for language in trans_list[1:]:
            trans_value = trans_before.get(language, {}).get(value, "")
            add_one_cell(row, col, trans_value)
            col += 1
        row += 1
    for word in need_trans:
        add_one_cell(row, 0, word)
        row += 1
    workbook.save(language_excel_path)


def update_front_json():
    result = {language: {} for language in trans_list}
    letter_list = string.uppercase
    if not os.path.exists(language_excel_path):
        print 'excel not exist'
        return 0
    workbook = load_workbook(language_excel_path)
    sheet = workbook.get_active_sheet()
    max_row = sheet.max_row
    row = 2
    for index in xrange(max_row-1):
        col = 0
        zh = trans_list[0]
        place = letter_list[col] + str(row)
        chinese = sheet[place].value
        col += 1
        result[zh][chinese] = chinese
        for language in trans_list[1:]:
            place = letter_list[col] + str(row)
            content = sheet[place].value
            if content:
                result[language][chinese] = content
            col += 1
        row += 1
    with open(language_json_path, "w+") as f:
        f.write(json.dumps(result))


if len(sys.argv) < 2:
    create_need_trans()
    print 'create excel success'
else:
    if sys.argv[1] == '1':
        update_front_json()
        print 'update front_json success'
    else:
        create_need_trans()
        print 'create excel success'








